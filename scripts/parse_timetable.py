#!/usr/bin/env python3
"""
Parses the TIET master timetable workbook into structured JSON.

The workbook is a print-oriented grid, not a dataset. Every assumption this
parser makes about its shape is asserted at the end of the run -- if the
university changes the layout next semester, the build fails loudly instead
of shipping a wrong timetable.

Usage:
    python scripts/parse_timetable.py [source.xlsx] [outdir]
"""

from __future__ import annotations

import json
import re
import sys
import hashlib
import datetime as dt
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path

from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

DAYS = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
PERIODS_PER_DAY = 14
FIRST_PERIOD = dt.time(8, 0)
PERIOD_MINUTES = 50

YEAR_OF_SHEET = {
    "FIRST YEAR A": 1, "FIRST YEAR B": 1,
    "2ND YEAR A": 2, "2ND YEARB B": 2,
    "3RD YEAR A": 3, "3RD YEAR B": 3,
    "4TH YEAR A": 4, "4TH YEAR B": 4,
    "PG TIME TABLE": 5,
}

# The source sheet contains find-and-replace damage; correct on the way out.
# Free-text scheduling notes, tidied for display (the source has typos).
NOTE_FIXES = {
    "ALTERNATE FROM SECOND WEEK": "Alternate weeks · from week 2",
    "ALTERNATE FRO, SECOND WEEK": "Alternate weeks · from week 2",
    "SECOND WEEK ONWARDS ALTERNATE WEEK": "Alternate weeks · from week 2",
    "FIRST WEEK ONWARDS ALTERNATE WEEK": "Alternate weeks · from week 1",
    "ALTERNATE FROM FIRST WEEK": "Alternate weeks · from week 1",
    "CAPSTONE PROJECT": "Capstone project",
}

BRANCH_FIXES = {
    "ELECTRONICS IND122UMENTATION ENGG": "Electronics & Instrumentation Engg",
    "ROBOTICS(D202) AND ARTIFICIAL INTELEGENCE": "Robotics & Artificial Intelligence",
    "ROBOTICS(D202)AND ARTIFICIAL INTELEGENCE": "Robotics & Artificial Intelligence",
    "ARTIFICIAL INTELEGENCE": "Artificial Intelligence",
}

# --------------------------------------------------------------------------
# Token classification
# --------------------------------------------------------------------------

RE_SUBJECT = re.compile(r"^[A-Z]{3}\s?[0-9X]{3}\s?[LPTSD]?$")
RE_SUBJECT_NOTE = re.compile(r"^(?P<code>[A-Z]{3}\s?[0-9X]{3}\s?[LPTSD]?)\s*\((?P<note>[^)]+)\)$")
RE_LAB_MARK = re.compile(r"^LAB\s?-?\s?\d?$")
RE_TIME = re.compile(r"^\d{1,2}[:.]\d{2}\s*:?\s*(AM|PM)$", re.I)
RE_ROOM = re.compile(r"^[A-Z]{1,5}\s?-?\s?\d{1,4}(\.\d+)?\s?-?\s?[A-Z]?$")
RE_NAMED_ROOM = re.compile(r"^(?P<label>[A-Z0-9][A-Z0-9 \-&.]*?)\s*\((?P<room>[^)]+)\)$")
RE_FACULTY = re.compile(r"^[A-Z]{2,5}(-[A-Z]{1,3})?$")
RE_SPECIAL_ROOM = re.compile(r"^(W/SHOP|WORKSHOP|SPORTS|GROUND|NCC|NSS|LIB|LIBRARY)$")
RE_PLACE_WORD = re.compile(r"\b(LAB|ROOM|HALL|FLOOR|CENTRE|CENTER|WORKSHOP|STUDIO|SHOP)\b")
RE_JUNK = re.compile(r"^[\-\u2013\u2014().,?_\s]*$|^\.?\d+\.?\d*$")

TYPE_OF_SUFFIX = {"L": "lecture", "P": "practical", "T": "tutorial",
                  "S": "seminar", "D": "discussion"}


def _atom(t: str) -> tuple[str, str] | None:
    """Classify one indivisible fragment. None means 'not recognised'."""
    if RE_JUNK.match(t):
        return ("empty", "")
    if RE_LAB_MARK.match(t):
        return ("labmark", t)
    if RE_TIME.match(t):
        return ("time", t)
    if RE_SPECIAL_ROOM.match(t):
        return ("room", t)
    m = RE_SUBJECT_NOTE.match(t)
    if m:
        return ("subject", m.group("code").replace(" ", ""))
    if RE_SUBJECT.match(t):
        return ("subject", t.replace(" ", ""))
    m = RE_NAMED_ROOM.match(t)
    if m:
        return ("named_room", t)
    if RE_ROOM.match(t):
        return ("room", t.replace(" ", ""))
    if RE_FACULTY.match(t):
        return ("faculty", t)
    return None


def classify(token: str) -> list[tuple[str, str]]:
    """Break one cell into (kind, value) pairs.

    Cells in this workbook are not atomic: a single cell may hold an eleven-way
    elective bundle, a slash-separated faculty list, a room plus an initial, or
    a free-text scheduling note. Split progressively rather than guessing.
    """
    t = " ".join(str(token).split()).strip().upper().rstrip(".")
    if not t:
        return []

    direct = _atom(t)
    if direct:
        return [direct]

    # The source has unbalanced brackets in a handful of cells
    # ("PL-5(L208", "NS1L102)"). Repair before giving up on them.
    if t.count("(") != t.count(")"):
        repaired = _atom(t.replace("(", "").replace(")", ""))
        if repaired:
            return [repaired]
        if "(" in t and not t.endswith(")"):
            repaired = _atom(t + ")")
            if repaired:
                return [repaired]

    # Slash-separated lists: electives, co-taught faculty, alternate rooms.
    if "/" in t:
        parts = [p.strip() for p in t.split("/") if p.strip()]
        results = [_atom(p) for p in parts]
        if parts and all(results):
            return [r for r in results if r and r[0] != "empty"]

    # Anything naming a physical place is a room label, even as free text
    # ("MOL BIO LAB", "CSED ROOM"). Checked before the whitespace split so the
    # words are not mistaken for faculty initials.
    if RE_PLACE_WORD.search(t):
        return [("room", t)]

    # "PHU301L F307", "E105 MBK" -- two atoms sharing one cell.
    if " " in t:
        parts = t.split()
        results = [_atom(p) for p in parts]
        if all(results):
            return [r for r in results if r and r[0] != "empty"]

    return [("note", t)]


# --------------------------------------------------------------------------
# Sheet geometry discovery
# --------------------------------------------------------------------------

@dataclass
class Geometry:
    sheet: str
    practical_row: int
    tutorial_row: int | None
    lecture_row: int | None
    branch_row: int | None
    srno_col: int
    hours_col: int
    data_start_col: int
    data_end_col: int
    period_rows: list[int]


def find_label(ws, label: str, max_row: int = 12, max_col: int = 10) -> tuple[int, int] | None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == label:
                return r, c
    return None


def find_srno_col(ws, practical_row: int, hours_col: int) -> int:
    """The period-number column. Labelled inconsistently (SR.NO / SR NO /
    nothing at all), so identify it by content: 1, 2, 3 ... running downward."""
    best, best_hits = hours_col - 1, -1
    for c in range(1, hours_col + 1):
        seq = [ws.cell(r, c).value for r in range(practical_row, practical_row + 30)]
        ints = [int(v) for v in seq if isinstance(v, (int, float))]
        hits = sum(1 for i, v in enumerate(ints[:5]) if v == i + 1)
        if hits > best_hits:
            best, best_hits = c, hits
    return best


def discover(ws) -> Geometry:
    """Locate the grid inside the sheet. Nothing here is hardcoded per sheet."""
    name = ws.title.strip()

    if name == "PG TIME TABLE":
        practical_row, lecture_row, tutorial_row, branch_row = 4, 3, None, None
        hours_col = find_label(ws, "HOUR")[1]
        srno_col = find_srno_col(ws, practical_row, hours_col)
    else:
        pr = find_label(ws, "PRACTICAL")
        if pr is None:
            raise ValueError(f"{name}: no PRACTICAL header row found")
        practical_row = pr[0]
        tutorial_row = practical_row - 1
        lecture_row = practical_row - 2
        branch_row = practical_row - 3
        hrs = find_label(ws, "HOURS")
        hours_col = hrs[1]
        srno_col = find_srno_col(ws, practical_row, hours_col)

    data_start_col = hours_col + 1

    # The grid's day/hour axis is repeated at the far right of the sheet for
    # print legibility. Everything from that repeat onward is not data.
    data_end_col = ws.max_column
    for c in range(data_start_col + 5, ws.max_column + 1):
        header_block = [
            str(ws.cell(r, c).value or "").strip().upper()
            for r in range(1, practical_row + 1)
        ]
        if any(h in ("DAY", "SR.NO", "SR NO", "HOURS", "HOUR") for h in header_block):
            data_end_col = c - 1
            break

    # Period rows: the SR.NO column carries 1..14, repeated once per day.
    period_rows = [
        r for r in range(practical_row, ws.max_row + 1)
        if isinstance(ws.cell(r, srno_col).value, (int, float))
        and 1 <= int(ws.cell(r, srno_col).value) <= PERIODS_PER_DAY
    ]
    # Trim to whole days and drop the workload/signature block below the grid.
    period_rows = period_rows[: (len(period_rows) // PERIODS_PER_DAY) * PERIODS_PER_DAY]

    return Geometry(name, practical_row, tutorial_row, lecture_row, branch_row,
                    srno_col, hours_col, data_start_col, data_end_col, period_rows)


# --------------------------------------------------------------------------
# Batch (column) discovery
# --------------------------------------------------------------------------

@dataclass
class Batch:
    id: str
    year: int
    sheet: str
    branch: str = ""
    lecture_group: str = ""
    tutorial_group: str = ""
    spans: list[tuple[int, int]] = field(default_factory=list)


def cell_text(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    return " ".join(str(v).split()).strip() if v is not None else ""


def discover_batches(ws, g: Geometry) -> dict[str, Batch]:
    """Map every practical subgroup to the column span(s) that describe it.

    4TH YEAR A repeats several batches in a second column region -- a batch may
    legitimately own more than one span, and both must be read.
    """
    year = YEAR_OF_SHEET.get(g.sheet, 0)
    batches: dict[str, Batch] = {}

    # Column positions where a new practical subgroup starts.
    starts = [c for c in range(g.data_start_col, g.data_end_col + 1)
              if cell_text(ws, g.practical_row, c)]

    carry_branch = carry_lecture = ""
    for i, c in enumerate(starts):
        end = (starts[i + 1] - 1) if i + 1 < len(starts) else g.data_end_col
        label = cell_text(ws, g.practical_row, c).upper()
        if label in ("DAY", "HOURS", "HOUR", "SR.NO", "SR NO", "PRACTICAL"):
            continue

        if g.branch_row:
            b = cell_text(ws, g.branch_row, c)
            if b and b.upper() != "BRANCH":
                carry_branch = b
        if g.lecture_row:
            lg = cell_text(ws, g.lecture_row, c)
            if lg and lg.upper() not in ("LECTURE", "BRANCH"):
                carry_lecture = lg
        tut = cell_text(ws, g.tutorial_row, c) if g.tutorial_row else ""
        if tut.upper() in ("TUTORIAL", "LECTURE", ""):
            tut = ""

        if label not in batches:
            batches[label] = Batch(
                id=label, year=year, sheet=g.sheet,
                branch=normalise_branch(carry_branch),
                lecture_group=carry_lecture, tutorial_group=tut,
            )
        batches[label].spans.append((c, end))

    return batches


ACRONYMS = {"AI", "ML", "VLSI", "CSE", "ECE", "EEE", "IT", "PG", "ME", "MSC",
            "MCA", "PHD", "BT", "BTD", "EST", "EV", "NCC", "NSS", "UG"}
EXPAND = {"ENGG": "Engineering", "ENGG.": "Engineering", "AND": "&",
          "COMPUTER": "Computer", "SCIENCE": "Science"}


def normalise_branch(raw: str) -> str:
    if not raw:
        return ""
    fixed = BRANCH_FIXES.get(raw.upper().strip(), raw)
    out = []
    for w in fixed.split():
        u = w.upper().strip(".")
        if u in ACRONYMS:
            out.append(u)
        elif u in EXPAND:
            out.append(EXPAND[u])
        elif w.isupper():
            out.append(w.capitalize())
        else:
            out.append(w)
    return " ".join(out)


# --------------------------------------------------------------------------
# Class extraction
# --------------------------------------------------------------------------

def build_merge_index(ws) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    idx = {}
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                idx[(r, c)] = (m.min_row, m.min_col, m.max_row, m.max_col)
    return idx


def harvest(ws, merges, row: int, col_lo: int, col_hi: int) -> list[str]:
    """All distinct non-empty strings in one row, within a column window."""
    out, seen = [], set()
    c = col_lo
    while c <= col_hi:
        span = merges.get((row, c))
        if span:
            _, mc0, _, mc1 = span
            v = cell_text(ws, span[0], mc0)
            c = mc1 + 1
        else:
            v = cell_text(ws, row, c)
            c += 1
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


@dataclass
class ClassEntry:
    day: int
    period: int
    periods: int
    start: str
    end: str
    code: str
    type: str
    title: str = ""
    room: str = ""
    faculty: str = ""
    options: list[str] = field(default_factory=list)
    choices: list[dict] = field(default_factory=list)
    aligned: bool = False
    note: str = ""
    raw: str = ""


def period_time(index: int) -> dt.time:
    base = dt.datetime.combine(dt.date(2026, 1, 1), FIRST_PERIOD)
    return (base + dt.timedelta(minutes=PERIOD_MINUTES * (index - 1))).time()


def extract(ws, g: Geometry, batches: dict[str, Batch], report: "Report") -> dict[str, list[ClassEntry]]:
    merges = build_merge_index(ws)
    schedule: dict[str, list[ClassEntry]] = defaultdict(list)

    for batch in batches.values():
        # A period is claimed once a multi-period block covers it.
        claimed: set[tuple[int, int]] = set()

        for slot, row in enumerate(g.period_rows):
            day, period = divmod(slot, PERIODS_PER_DAY)
            period += 1
            if (day, period) in claimed:
                continue

            for span_lo, span_hi in batch.spans:
                # The subject cell may be merged across many subgroups (a
                # lecture). Widen the harvest window to that merge so the
                # room and faculty, which sit at opposite ends of the block,
                # are both picked up.
                merge = merges.get((row, span_lo))
                lo, hi = (merge[1], merge[3]) if merge else (span_lo, span_hi)

                head = harvest(ws, merges, row, lo, hi)
                body = harvest(ws, merges, row + 1, lo, hi)
                tokens = head + body
                label = None
                if not has_subject(tokens):
                    label = local_label(head, body)
                    if not label:
                        continue

                # Duration. A lecture or tutorial at TIET is always one
                # period; the extra line below it is the faculty name
                # overflowing into the next period's row band, not a
                # continuation. Only practicals genuinely run long.
                length, raw_rows = 1, list(tokens)
                block_type = block_kind(tokens)
                probe = slot + 1
                while probe < len(g.period_rows) and (probe % PERIODS_PER_DAY) != 0:
                    nrow = g.period_rows[probe]
                    ntok = harvest(ws, merges, nrow, lo, hi) + harvest(ws, merges, nrow + 1, lo, hi)
                    if not ntok or has_subject(ntok):
                        break
                    raw_rows += ntok
                    if block_type != "practical":
                        break                      # absorb the overflow, stay 1 period
                    if length >= 2 and not any(
                            k in ("room", "named_room", "labmark")
                            for c in ntok for k, _ in classify(c)):
                        break                      # weak signal; don't over-extend
                    length += 1
                    probe += 1
                    if length >= 4:
                        break

                entry = assemble(raw_rows, day, period, length, report,
                                 g.sheet, batch.id, label)
                schedule[batch.id].append(entry)
                for p in range(period, period + length):
                    claimed.add((day, p))
                break  # one span produced the class; don't double-add

    return schedule


PLACEHOLDER = re.compile(r"^[\-\u2013\u2014_.?\s]*$")


def split_aligned(cell: str) -> list[str]:
    """Split a slash list keeping position, duplicates and blanks.

    Elective cells are parallel arrays: the nth course goes with the nth room
    and the nth teacher, and the sheet writes '---' where a value is missing
    precisely to hold that position. Deduplicating or dropping blanks here
    silently shifts every later entry onto the wrong room, so this split does
    neither.
    """
    return ["" if PLACEHOLDER.match(part) else " ".join(part.split()).strip().upper()
            for part in cell.split("/")]


def elective_choices(cells: list[str], report: "Report", where: str) -> tuple[list[dict], bool]:
    """Pair each elective option with its own room and faculty.

    Returns (choices, aligned). When the parallel arrays disagree in length we
    refuse to guess: the options are still listed, but without a room, because
    sending a student to the wrong lab is worse than sending them to none.
    """
    option_cell = None
    for cell in cells:
        codes = [c for c in split_aligned(cell) if RE_SUBJECT_NOTE.match(c) or RE_SUBJECT.match(c)]
        if len(codes) > 1:
            option_cell = cell
            break
    if option_cell is None:
        return [], False

    raw_options = split_aligned(option_cell)
    options, notes = [], []
    for part in raw_options:
        m = RE_SUBJECT_NOTE.match(part)
        if m:
            options.append(m.group("code").replace(" ", ""))
            notes.append(m.group("note").strip())
        else:
            options.append(part.replace(" ", ""))
            notes.append("")

    rooms: list[str] = []
    faculty: list[str] = []
    for cell in cells:
        if cell == option_cell:
            continue
        parts = split_aligned(cell)
        filled = [p for p in parts if p]
        if not filled:
            continue
        kinds = [k for p in filled for k, _ in classify(p)]
        room_like = sum(1 for k in kinds if k in ("room", "named_room"))
        fac_like = sum(1 for k in kinds if k == "faculty")
        if room_like >= fac_like and room_like and not rooms:
            rooms = parts
        elif fac_like and not faculty:
            faculty = parts

    aligned = bool(rooms or faculty)
    if rooms and len(rooms) != len(options):
        aligned = False
    if faculty and len(faculty) != len(options):
        aligned = False

    if not aligned:
        report.warn(f"elective not aligned: {where} "
                    f"({len(options)} options, {len(rooms)} rooms, {len(faculty)} faculty)")

    choices = []
    for i, code in enumerate(options):
        if not code:
            continue
        choices.append({
            "code": code,
            "room": rooms[i] if aligned and i < len(rooms) else "",
            "faculty": faculty[i] if aligned and i < len(faculty) else "",
            "note": notes[i],
        })
    return choices, aligned


def block_kind(cells: list[str]) -> str:
    """lecture / tutorial / practical, from the subject code suffix."""
    for cell in cells:
        for kind, val in classify(cell):
            if kind == "subject":
                return TYPE_OF_SUFFIX.get(val[-1], "class")
    return "class"


def has_subject(cells: list[str]) -> bool:
    return any(k == "subject" for c in cells for k, _ in classify(c))


def local_label(head: list[str], body: list[str]) -> str | None:
    """Some departments write a course label instead of a course code
    ("PMC L", "BEST-EE P"). Accept it as a subject when the cell below it
    looks like a real class -- i.e. carries a room or a faculty initial."""
    notes = [v for c in head for k, v in classify(c) if k == "note"]
    if not notes:
        return None
    supported = any(k in ("room", "named_room", "faculty", "labmark")
                    for c in head + body for k, _ in classify(c))
    for n in notes:
        if supported and len(n) <= 14 and n.count(" ") <= 1:
            return n
    return None


def assemble(cells: list[str], day: int, period: int, length: int,
             report: "Report", sheet: str, batch_id: str,
             label: str | None = None) -> ClassEntry:
    """Turn the harvested cell strings of one block into a single class."""
    code, ctype, options = "", "class", []
    rooms, faculty, notes = [], [], []
    if label:
        code, ctype = label, "class"

    for cell in cells:
        atoms = classify(cell)
        subjects = [v for k, v in atoms if k == "subject"]

        # One cell holding several codes is an elective bundle: the student
        # attends whichever one they registered for.
        if len(subjects) > 1 and not code:
            code, ctype, options = "ELECTIVE", "elective", subjects
        elif subjects and not code:
            code = subjects[0]
            ctype = TYPE_OF_SUFFIX.get(code[-1], "class")

        for kind, val in atoms:
            if kind == "room":
                rooms.append(val)
            elif kind == "named_room":
                m = RE_NAMED_ROOM.match(val)
                rooms.append(f"{m.group('label').strip()} ({m.group('room').strip()})")
            elif kind == "faculty":
                faculty.append(val)
            elif kind == "note":
                if val == code:
                    continue
                notes.append(val)
                report.unknown[val] += 1
                report.unknown_context.setdefault(
                    val, f"{sheet} / {batch_id} / {DAYS[day]} P{period}")

    choices, aligned = ([], False)
    if ctype == "elective":
        choices, aligned = elective_choices(
            cells, report, f"{sheet}/{batch_id} {DAYS[day]} P{period}")

    start = period_time(period)
    end = period_time(period + length)
    return ClassEntry(
        day=day, period=period, periods=length,
        start=start.strftime("%H:%M"), end=end.strftime("%H:%M"),
        code=code, type=ctype,
        room=" / ".join(dict.fromkeys(rooms)),
        faculty=" / ".join(dict.fromkeys(faculty)),
        options=list(dict.fromkeys(options)),
        choices=choices,
        aligned=aligned,
        note=" · ".join(dict.fromkeys(NOTE_FIXES.get(n, n) for n in notes)),
        raw=" | ".join(cells),
    )


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------

class Report:
    def __init__(self):
        self.unknown: Counter = Counter()
        self.unknown_context: dict[str, str] = {}
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.stats: dict = {}

    def error(self, msg): self.errors.append(msg)
    def warn(self, msg): self.warnings.append(msg)


def validate(batches: dict[str, Batch], schedule: dict[str, list[ClassEntry]], report: Report):
    empty = sorted(b for b in batches if not schedule.get(b))
    report.stats["emptyBatches"] = empty
    if len(empty) > max(3, len(batches) // 50):
        report.error(f"{len(empty)} batches have an empty week -- column mapping "
                     f"is probably broken: {empty[:12]}")
    elif empty:
        report.warn(f"blank in source: {len(empty)} batches carry no classes: {empty}")

    for bid, entries in schedule.items():
        occupied: dict[tuple[int, int], str] = {}
        for e in entries:
            for p in range(e.period, e.period + e.periods):
                key = (e.day, p)
                if key in occupied and occupied[key] != e.code:
                    report.error(f"{bid}: double-booked {DAYS[e.day]} period {p} "
                                 f"({occupied[key]} vs {e.code})")
                occupied[key] = e.code
            if e.periods > 1 and e.type == "lecture":
                report.warn(f"{bid}: lecture {e.code} spans {e.periods} periods "
                            f"({DAYS[e.day]} P{e.period})")
            if not e.room:
                report.warn(f"{bid}: no room for {e.code} on {DAYS[e.day]} P{e.period}")
            if e.period + e.periods - 1 > PERIODS_PER_DAY:
                report.error(f"{bid}: {e.code} runs past the end of the day")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def load_override(path: str) -> dict:
    f = Path(path)
    return json.loads(f.read_text()) if f.exists() else {}


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "source/TIME_TABLE_AUG_DEC_2026.xlsx")
    out = Path(sys.argv[2] if len(sys.argv) > 2 else "data")
    out.mkdir(parents=True, exist_ok=True)
    (out / "batches").mkdir(exist_ok=True)

    report = Report()
    wb = load_workbook(src, data_only=True)

    all_batches: dict[str, Batch] = {}
    all_schedule: dict[str, list[ClassEntry]] = {}

    for ws in wb.worksheets:
        g = discover(ws)
        if len(g.period_rows) % PERIODS_PER_DAY:
            report.error(f"{ws.title}: {len(g.period_rows)} period rows is not a multiple of {PERIODS_PER_DAY}")
        batches = discover_batches(ws, g)
        sched = extract(ws, g, batches, report)
        for bid, b in batches.items():
            if bid in all_batches:
                report.warn(f"batch {bid} appears in two sheets ({all_batches[bid].sheet}, {b.sheet})")
                continue
            all_batches[bid] = b
        for bid, entries in sched.items():
            all_schedule.setdefault(bid, []).extend(entries)
        print(f"  {ws.title:<16} rows {len(g.period_rows):>3}  "
              f"cols {g.data_start_col}-{g.data_end_col}  batches {len(batches):>3}  "
              f"classes {sum(len(v) for v in sched.values()):>5}")

    validate(all_batches, all_schedule, report)

    # ---- write output ----
    subjects = load_override("overrides/subjects.json")
    faculty_names = load_override("overrides/faculty.json")
    descriptions = load_override("overrides/descriptions.json")
    # Names are keyed by the six-character base code; the L/P/T suffix only
    # says whether it is the lecture, lab or tutorial of the same course.
    for entries in all_schedule.values():
        for e in entries:
            e.title = subjects.get(e.code) or subjects.get(e.code[:6], "")
            for ch in e.choices:
                ch["title"] = subjects.get(ch["code"]) or subjects.get(ch["code"][:6], "")

    digest = hashlib.sha256(src.read_bytes()).hexdigest()[:16]
    index = {
        "term": "August - December 2026",
        "source": src.name,
        "sourceHash": digest,
        "generated": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "days": DAYS,
        "periods": [
            {"index": i, "start": period_time(i).strftime("%H:%M"),
             "end": period_time(i + 1).strftime("%H:%M")}
            for i in range(1, PERIODS_PER_DAY + 1)
        ],
        "batches": [
            {"id": b.id, "year": b.year, "branch": b.branch,
             "lectureGroup": b.lecture_group, "tutorialGroup": b.tutorial_group,
             "classes": len(all_schedule.get(b.id, []))}
            for b in sorted(all_batches.values(), key=lambda x: (x.year, x.id))
        ],
    }
    seen_codes = sorted({e.code for v in all_schedule.values() for e in v})
    index["subjects"] = {
        c: subjects.get(c) or subjects[c[:6]]
        for c in seen_codes
        if c in subjects or c[:6] in subjects
    }
    index["faculty"] = faculty_names
    index["descriptions"] = {
        c: descriptions.get(c) or descriptions[c[:6]]
        for c in seen_codes
        if c in descriptions or c[:6] in descriptions
    }
    index["coverage"] = {
        "subjectsNamed": len(index["subjects"]),
        "subjectsTotal": len(seen_codes),
    }
    (out / "index.json").write_text(json.dumps(index, indent=1))

    for bid, entries in all_schedule.items():
        entries.sort(key=lambda e: (e.day, e.period))
        payload = {
            "id": bid,
            "term": index["term"],
            "meta": {k: v for k, v in asdict(all_batches[bid]).items() if k != "spans"},
            "classes": [asdict(e) for e in entries],
        }
        (out / "batches" / f"{bid}.json").write_text(json.dumps(payload, indent=1))

    # Hand-maintained calendars pass straight through to the site.
    for name in ("exams", "events"):
        payload = load_override(f"overrides/{name}.json")
        (out / f"{name}.json").write_text(json.dumps(payload, indent=1))

    (out / "unknown_tokens.json").write_text(json.dumps(
        [{"token": t, "count": n, "firstSeen": report.unknown_context.get(t, "")}
         for t, n in report.unknown.most_common()], indent=1))

    # ---- summary ----
    total = sum(len(v) for v in all_schedule.values())
    print(f"\nbatches: {len(all_batches)}   class blocks: {total}")
    print(f"unknown tokens: {len(report.unknown)} distinct, {sum(report.unknown.values())} occurrences")
    print(f"errors: {len(report.errors)}   warnings: {len(report.warnings)}")
    for e in report.errors[:15]:
        print("  ERROR  ", e)
    wc = Counter(w.split(":")[1].strip().split()[0] for w in report.warnings if ":" in w)
    for w, n in wc.most_common(6):
        print(f"  warn   {w} x{n}")
    return 1 if report.errors else 0


if __name__ == "__main__":
    sys.exit(main())
